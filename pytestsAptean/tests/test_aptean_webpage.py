import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, ElementClickInterceptedException, \
    StaleElementReferenceException
import time
import os
from urllib.parse import urljoin, urlparse
import json


class ApteanInteractiveNavScraper:
    def __init__(self):
        self.all_links = []
        self.nav_structure = {}
        self.screenshots_dir = 'aptean_screenshots'
        self.driver = None
        self.base_url = 'https://www.aptean.com'
        self.actions = None

        # Create screenshots directory
        if not os.path.exists(self.screenshots_dir):
            os.makedirs(self.screenshots_dir)

        self.setup_selenium()

    def setup_selenium(self):
        """Setup Selenium WebDriver with optimal settings"""
        chrome_options = Options()
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--window-size=1920,1080")
        chrome_options.add_argument("--disable-extensions")
        chrome_options.add_argument("--start-maximized")

        # Optional: Run headless (comment out to see browser)
        # chrome_options.add_argument("--headless")

        try:
            self.driver = webdriver.Chrome(options=chrome_options)
            self.driver.set_page_load_timeout(30)
            self.actions = ActionChains(self.driver)
        except Exception as e:
            print(f"Error setting up Chrome driver: {e}")
            print("Please ensure ChromeDriver is installed and in PATH")

    def scrape_interactive_nav(self):
        """Click on every button/dropdown in navigation and capture all links"""
        print(f"Starting interactive navigation scrape of {self.base_url}")

        try:
            # Load the main page
            self.driver.get(self.base_url)
            time.sleep(3)

            # Handle cookie popup
            self._handle_popup()

            # Find navigation container
            nav_container = self._find_nav_container()
            if not nav_container:
                print("Could not find navigation container")
                return

            # Get all clickable elements in navigation
            clickable_elements = self._find_clickable_nav_elements(nav_container)
            print(f"Found {len(clickable_elements)} clickable navigation elements")

            # Process each clickable element
            for i, element_info in enumerate(clickable_elements, 1):
                print(f"\n[{i}/{len(clickable_elements)}] Processing: {element_info['text']}")
                self._process_nav_element(element_info, i)

                # Brief pause between interactions
                time.sleep(1)

            print(f"\nCompleted interactive scraping. Found {len(self.all_links)} total links.")
            return len(self.all_links)

        except Exception as e:
            print(f"Error during interactive scraping: {e}")
            return 0

    def _handle_popup(self):
        """Handle cookie/popup dialogs"""
        popup_selectors = [
            "button[contains(text(), 'Accept')]",
            "button[contains(text(), 'OK')]",
            "button[contains(text(), 'Close')]",
            ".cookie-accept",
            ".accept-cookies",
            "#accept-cookies",
            ".modal-close",
            "[aria-label*='close']",
            "[aria-label*='Close']"
        ]

        for selector in popup_selectors:
            try:
                popup = WebDriverWait(self.driver, 2).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, selector))
                )
                popup.click()
                print("✓ Closed popup dialog")
                time.sleep(1)
                break
            except:
                continue

    def _find_nav_container(self):
        """Find the main navigation container"""
        nav_selectors = [
            "nav",
            ".main-nav",
            ".navigation",
            ".navbar",
            ".header-nav",
            ".primary-nav",
            ".top-nav",
            "[role='navigation']",
            ".menu-container"
        ]

        for selector in nav_selectors:
            try:
                nav = self.driver.find_element(By.CSS_SELECTOR, selector)
                if nav.is_displayed():
                    print(f"✓ Found navigation container: {selector}")
                    return nav
            except:
                continue

        return None

    def _find_clickable_nav_elements(self, nav_container):
        """Find all clickable elements in navigation"""
        clickable_elements = []

        # Selectors for different types of clickable elements
        clickable_selectors = [
            "button",
            "a[href]",
            "[role='button']",
            ".dropdown-toggle",
            ".menu-toggle",
            "[data-toggle]",
            "[aria-expanded]",
            ".has-dropdown",
            ".menu-item-has-children",
            "li:has(ul)",
            "[onclick]"
        ]

        for selector in clickable_selectors:
            try:
                elements = nav_container.find_elements(By.CSS_SELECTOR, selector)
                for element in elements:
                    if element.is_displayed() and element.is_enabled():
                        text = element.text.strip()
                        tag_name = element.tag_name
                        classes = element.get_attribute('class') or ''

                        # Skip if already processed or empty text
                        if not text or len(text) > 100:
                            continue

                        element_info = {
                            'element': element,
                            'text': text,
                            'tag_name': tag_name,
                            'classes': classes,
                            'selector': selector,
                            'href': element.get_attribute('href'),
                            'has_dropdown': 'dropdown' in classes.lower() or
                                            element.get_attribute('aria-expanded') is not None
                        }

                        # Avoid duplicates
                        if not any(e['text'] == text for e in clickable_elements):
                            clickable_elements.append(element_info)

            except Exception as e:
                continue

        return clickable_elements

    def _process_nav_element(self, element_info, index):
        """Process a single navigation element"""
        element = element_info['element']
        text = element_info['text']

        try:
            # Record the element itself if it's a link
            if element_info['href']:
                self._add_link_data(text, element_info['href'], 'direct_link', index)

            # Try different interaction methods
            interaction_success = False

            # Method 1: Regular click
            if not interaction_success:
                interaction_success = self._try_click(element, 'click')

            # Method 2: Hover (for hover dropdowns)
            if not interaction_success:
                interaction_success = self._try_hover(element, 'hover')

            # Method 3: JavaScript click
            if not interaction_success:
                interaction_success = self._try_js_click(element, 'js_click')

            if interaction_success:
                # Wait for dropdown/submenu to appear
                time.sleep(1)

                # Capture links that appeared
                self._capture_revealed_links(text, index)

                # Take screenshot of the opened dropdown
                self._take_dropdown_screenshot(text, index)

                # Close dropdown if possible
                self._close_dropdown()

        except Exception as e:
            print(f"Error processing {text}: {e}")

    def _try_click(self, element, method_name):
        """Try regular click"""
        try:
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
            time.sleep(0.5)
            element.click()
            print(f"✓ {method_name} successful")
            return True
        except Exception as e:
            return False

    def _try_hover(self, element, method_name):
        """Try hover interaction"""
        try:
            self.actions.move_to_element(element).perform()
            time.sleep(1)
            print(f"✓ {method_name} successful")
            return True
        except Exception as e:
            return False

    def _try_js_click(self, element, method_name):
        """Try JavaScript click"""
        try:
            self.driver.execute_script("arguments[0].click();", element)
            print(f"✓ {method_name} successful")
            return True
        except Exception as e:
            return False

    def _capture_revealed_links(self, parent_text, parent_index):
        """Capture links that were revealed by interaction"""
        try:
            # Look for newly visible links
            new_link_selectors = [
                ".dropdown-menu a",
                ".submenu a",
                ".mega-menu a",
                "[aria-expanded='true'] a",
                ".show a",
                ".open a",
                ".active a",
                "ul li a[href]"
            ]

            for selector in new_link_selectors:
                try:
                    links = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    for link in links:
                        if link.is_displayed():
                            href = link.get_attribute('href')
                            text = link.text.strip()

                            if href and text and self._is_valid_link(href):
                                self._add_link_data(
                                    text,
                                    href,
                                    f'dropdown_from_{parent_text}',
                                    parent_index,
                                    parent_text
                                )
                except:
                    continue

        except Exception as e:
            print(f"Error capturing revealed links: {e}")

    def _add_link_data(self, text, url, source_type, parent_index, parent_text=''):
        """Add link data to collection"""
        # Avoid duplicates
        if any(link['url'] == url and link['text'] == text for link in self.all_links):
            return

        link_data = {
            'text': text,
            'url': url,
            'source_type': source_type,
            'parent_element': parent_text,
            'parent_index': parent_index,
            'domain': urlparse(url).netloc,
            'is_external': not url.startswith(self.base_url),
            'screenshot_path': '',
            'screenshot_status': 'pending',
            'page_title': '',
            'scraped_at': pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')
        }

        self.all_links.append(link_data)
        print(f"  📎 Added link: {text} -> {url}")

    def _take_dropdown_screenshot(self, element_text, index):
        """Take screenshot of opened dropdown"""
        try:
            safe_name = self._sanitize_filename(f"{index:03d}_dropdown_{element_text}")
            screenshot_path = os.path.join(self.screenshots_dir, f"{safe_name}.png")
            self.driver.save_screenshot(screenshot_path)
            print(f"  📸 Dropdown screenshot: {screenshot_path}")
        except Exception as e:
            print(f"  ❌ Screenshot failed: {e}")

    def _close_dropdown(self):
        """Try to close any open dropdowns"""
        try:
            # Click somewhere neutral to close dropdowns
            body = self.driver.find_element(By.TAG_NAME, 'body')
            self.actions.move_to_element_with_offset(body, 10, 10).click().perform()
            time.sleep(0.5)
        except:
            pass

        # Try pressing escape
        try:
            self.driver.find_element(By.TAG_NAME, 'body').send_keys('\ue000')  # ESC key
        except:
            pass

    def _is_valid_link(self, href):
        """Check if link is valid for processing"""
        if not href:
            return False

        skip_patterns = [
            'mailto:', 'tel:', 'javascript:', '#', 'void(0)',
            '.pdf', '.doc', '.zip', '.exe', '.jpg', '.png'
        ]

        return not any(pattern in href.lower() for pattern in skip_patterns)

    def take_page_screenshots(self, limit=None):
        """Take screenshots of all discovered pages"""
        if not self.all_links:
            print("No links found. Run scrape_interactive_nav() first.")
            return

        links_to_process = self.all_links[:limit] if limit else self.all_links
        successful_screenshots = 0

        for i, link_data in enumerate(links_to_process, 1):
            if link_data['source_type'] == 'dropdown_from_*':
                continue  # Skip dropdown screenshots

            url = link_data['url']
            text = link_data['text']

            print(f"[{i}/{len(links_to_process)}] Screenshotting: {text}")

            try:
                self.driver.get(url)
                WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.TAG_NAME, "body"))
                )
                time.sleep(2)

                # Get page title
                link_data['page_title'] = self.driver.title

                # Take screenshot
                safe_name = self._sanitize_filename(f"page_{i:03d}_{text}")
                screenshot_path = os.path.join(self.screenshots_dir, f"{safe_name}.png")
                self.driver.save_screenshot(screenshot_path)

                link_data['screenshot_path'] = screenshot_path
                link_data['screenshot_status'] = 'success'
                successful_screenshots += 1

                print(f"✓ Screenshot saved")

            except Exception as e:
                print(f"✗ Error: {e}")
                link_data['screenshot_status'] = f'error: {str(e)[:100]}'

            time.sleep(1)

        return successful_screenshots

    def export_to_excel(self, filename='aptean_interactive_nav_analysis.xlsx'):
        """Export all data to Excel"""
        if not self.all_links:
            print("No data to export!")
            return None

        df = pd.DataFrame(self.all_links)
        wb = Workbook()
        ws = wb.active
        ws.title = "Interactive Nav Links"

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        direct_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        dropdown_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        external_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        # Headers
        headers = ['Link Text', 'URL', 'Source Type', 'Parent Element', 'Domain',
                   'Is External', 'Page Title', 'Screenshot Path', 'Screenshot Status', 'Scraped At']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Data
        for row, data in enumerate(self.all_links, 2):
            ws.cell(row=row, column=1, value=data.get('text', ''))
            ws.cell(row=row, column=2, value=data.get('url', ''))
            ws.cell(row=row, column=3, value=data.get('source_type', ''))
            ws.cell(row=row, column=4, value=data.get('parent_element', ''))
            ws.cell(row=row, column=5, value=data.get('domain', ''))
            ws.cell(row=row, column=6, value=data.get('is_external', ''))
            ws.cell(row=row, column=7, value=data.get('page_title', ''))
            ws.cell(row=row, column=8, value=data.get('screenshot_path', ''))
            ws.cell(row=row, column=9, value=data.get('screenshot_status', ''))
            ws.cell(row=row, column=10, value=data.get('scraped_at', ''))

            # Color coding
            for col in range(1, 11):
                cell = ws.cell(row=row, column=col)
                cell.border = border

                if data.get('source_type') == 'direct_link':
                    cell.fill = direct_fill
                elif 'dropdown' in str(data.get('source_type', '')):
                    cell.fill = dropdown_fill
                elif data.get('is_external'):
                    cell.fill = external_fill

        # Column widths
        column_widths = [25, 60, 20, 25, 20, 12, 40, 50, 20, 18]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

        # Summary sheet
        summary_ws = wb.create_sheet("Summary")
        summary_data = self._generate_summary()

        for row, (key, value) in enumerate(summary_data.items(), 1):
            summary_ws.cell(row=row, column=1, value=key).font = Font(bold=True)
            summary_ws.cell(row=row, column=2, value=str(value))

        wb.save(filename)
        print(f"Excel file '{filename}' created successfully!")
        return filename

    def _generate_summary(self):
        """Generate summary statistics"""
        df = pd.DataFrame(self.all_links)

        direct_links = len(df[df['source_type'] == 'direct_link'])
        dropdown_links = len(df[df['source_type'].str.contains('dropdown', na=False)])

        return {
            'Total Links Found': len(self.all_links),
            'Direct Navigation Links': direct_links,
            'Dropdown/Submenu Links': dropdown_links,
            'External Links': df['is_external'].sum() if len(df) > 0 else 0,
            'Internal Links': len(df) - df['is_external'].sum() if len(df) > 0 else 0,
            'Screenshots Directory': self.screenshots_dir,
            'Analysis Date': pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')
        }

    def _sanitize_filename(self, filename):
        """Sanitize filename for saving"""
        invalid_chars = '<>:"/\\|?*'
        for char in invalid_chars:
            filename = filename.replace(char, '_')
        return filename[:100]

    def close(self):
        """Clean up resources"""
        if self.driver:
            self.driver.quit()
            print("Browser closed.")


def test_aptean_main_page():
    """Main execution function"""
    scraper = ApteanInteractiveNavScraper()

    try:
        print("=== Aptean Interactive Navigation Scraper ===")

        # Step 1: Interactive navigation scraping
        print("\n1. Clicking through all navigation elements...")
        link_count = scraper.scrape_interactive_nav()

        if link_count == 0:
            print("No links found!")
            return

        # Display summary
        print(f"\n📊 Discovery Summary:")
        direct_links = len([l for l in scraper.all_links if l['source_type'] == 'direct_link'])
        dropdown_links = len([l for l in scraper.all_links if 'dropdown' in l['source_type']])

        print(f"   Direct navigation links: {direct_links}")
        print(f"   Dropdown/submenu links: {dropdown_links}")
        print(f"   Total unique links: {link_count}")

        # Step 2: Screenshots
        response = input(f"\nTake screenshots of all {link_count} pages? (y/n/number): ").lower()

        if response == 'y':
            print("\n2. Taking page screenshots...")
            scraper.take_page_screenshots()
        elif response.isdigit():
            limit = int(response)
            print(f"\n2. Taking screenshots of first {limit} pages...")
            scraper.take_page_screenshots(limit=limit)

        # Step 3: Export
        print("\n3. Exporting to Excel...")
        filename = scraper.export_to_excel()

        print(f"\n🎉 Analysis Complete!")
        print(f"📊 Excel file: {filename}")
        print(f"📁 Screenshots: {scraper.screenshots_dir}/")
        print(f"🔗 Total links: {len(scraper.all_links)}")

    except KeyboardInterrupt:
        print("\n⚠️  Operation cancelled by user.")
    except Exception as e:
        print(f"❌ Error: {e}")
    finally:
        scraper.close()


# if __name__ == "__main__":
#     main()